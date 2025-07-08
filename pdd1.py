import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook

# === 1. Create a new Excel file ===
wb = Workbook()
ws = wb.active
ws.title = "MySheet"

# Add some data
ws['A1'] = "Name"
ws['B1'] = "Score"
ws.append(["Alice", 90])
ws.append(["Bob", 85])

# Save it
wb.save("sample.xlsx")
print("Excel file 'sample.xlsx' created successfully.")

# === 2. Load and Read the Excel file ===
wb2 = load_workbook("sample.xlsx")
ws2 = wb2.active

print("\nReading from 'sample.xlsx':")
for row in ws2.iter_rows(min_row=1, max_col=2, max_row=3, values_only=True):
    print(row)

st.set_page_config(page_title="Bacterial Secretome Web Tool", layout="wide")

# Load the dataset with error handling
@st.cache_data
def load_data():
    try:
        df = pd.read_excel("Bacterial_Secretome_Database_2600.xlsx", engine="openpyxl")
        return df
    except FileNotFoundError:
        st.error("❌ Dataset not found! Please ensure the Excel file is in the same folder as this app.")
        return pd.DataFrame()
    except Exception as e:
        st.error(f"⚠️ Error loading the dataset: {e}")
        return pd.DataFrame()

df = load_data()

# Create tabs
tab1, tab2, tab3, tab4 = st.tabs([
    "🦠 About Disease",
    "🔍 Search Protein",
    "🧬 GC Content Tool",
    "📊 View Database"
])

# --- Tab 1: Disease Info ---
with tab1:
    st.title("Bacterial Infections & Secretome Proteins")
    st.markdown("""
    ### 🔬 Overview
    Bacterial infections are caused by pathogenic bacteria and can affect various parts of the body.

    ### 🧫 Types
    - **Respiratory**: Tuberculosis, Pneumonia
    - **Digestive**: Cholera, Salmonella
    - **Skin**: Cellulitis, Impetigo

    ### 💊 Treatments
    - Antibiotics (Penicillin, Tetracycline)
    - Immunotherapy
    - Vaccination
    - Supportive care and hygiene

    Secretome proteins are exported by bacteria to interact with the host and enhance infection.
    """)

# --- Tab 2: Search Protein ---
with tab2:
    st.title("🔍 Protein Search Tool")
    if df.empty:
        st.warning("Load the dataset to use this feature.")
    else:
        search_input = st.text_input("Enter Protein Name (partial/full):")
        if search_input:
            results = df[df['Protein Name'].str.contains(search_input, case=False, na=False)]
            if not results.empty:
                st.success(f"✅ Found {len(results)} matching protein(s).")
                st.dataframe(results)
            else:
                st.warning("No protein matched your search.")

# --- Tab 3: GC Content Tool ---
with tab3:
    st.title("🧬 GC Content Calculator")
    dna_seq = st.text_area("Paste a DNA sequence here (A, T, G, C only):")
    if dna_seq:
        dna_seq = dna_seq.upper()
        valid_chars = {'A', 'T', 'G', 'C'}
        if all(base in valid_chars for base in dna_seq):
            gc_count = dna_seq.count('G') + dna_seq.count('C')
            gc_percent = (gc_count / len(dna_seq)) * 100
            st.success(f"GC Content: **{gc_percent:.2f}%**")
        else:
            st.error("Invalid characters detected! Use only A, T, G, C.")

# --- Tab 4: Database ---
with tab4:
    st.title("📊 Secretome Protein Database")
    if df.empty:
        st.warning("Database not loaded.")
    else:
        st.dataframe(df)
